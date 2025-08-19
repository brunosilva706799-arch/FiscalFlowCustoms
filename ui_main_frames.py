import os
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from collections import defaultdict
from PIL import Image, ImageTk, ImageEnhance
import ttkbootstrap as ttk
from datetime import datetime
import logging
from openpyxl import Workbook

# Importa as funções de lógica do nosso outro arquivo
import core_logic

# --- TELA DE NOVIDADES DA VERSÃO ---
class WhatsNewFrame(ttk.Frame):
    # Este __init__ permanece o mesmo, pois ele USA os argumentos
    def __init__(self, parent, controller, current_version, release_notes):
        super().__init__(parent)
        self.controller = controller
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)
        ttk.Label(self, text=f"Novidades da Versão {current_version}", font=("Helvetica", 22, "bold")).grid(row=0, column=0, pady=20)
        text_area = scrolledtext.ScrolledText(self, wrap=tk.WORD, height=10, width=50, font=("Helvetica", 10))
        text_area.grid(row=1, column=0, sticky="nsew", padx=50, pady=10)
        text_area.insert(tk.INSERT, release_notes)
        text_area.config(state='disabled')
        continue_button = ttk.Button(self, text="Entendido, continuar para o programa",
                                      command=self.close_and_continue, bootstyle="primary")
        continue_button.grid(row=2, column=0, pady=20)

    def close_and_continue(self):
        self.controller.save_config()
        self.controller.show_frame("HomeFrame")

# --- TELA DE INÍCIO (DASHBOARD) ---
class HomeFrame(ttk.Frame):
    # --- CORREÇÃO AQUI: Adicionado *args para aceitar argumentos extras ---
    def __init__(self, parent, controller, *args): 
        super().__init__(parent)
        self.controller = controller
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)
        self.rowconfigure(1, weight=1)
        title_frame = ttk.Frame(self)
        title_frame.grid(row=0, column=0, columnspan=2, pady=(40, 10), sticky='ew')
        title_frame.columnconfigure(0, weight=1)
        ttk.Label(title_frame, text="Fiscal Flow - Customs Trading", font=("Helvetica", 24, "bold")).pack()
        left_container = ttk.Frame(self)
        left_container.grid(row=1, column=0, sticky='nsew')
        buttons_frame = ttk.Frame(left_container)
        buttons_frame.pack(expand=True)
        style = self.controller.app_style 
        style.configure('Large.TButton', font=('Helvetica', 12))
        btn_contabil = ttk.Button(buttons_frame, text="Departamento Contábil", style='Large.TButton', width=40, state="disabled", bootstyle="secondary")
        btn_contabil.pack(pady=10, ipady=5)
        btn_fiscal = ttk.Button(buttons_frame, text="Departamento Fiscal", style='Large.TButton', width=40, state="disabled", bootstyle="secondary")
        btn_fiscal.pack(pady=10, ipady=5)
        btn_pessoal = ttk.Button(buttons_frame, text="Departamento Pessoal", style='Large.TButton', width=40, state="disabled", bootstyle="secondary")
        btn_pessoal.pack(pady=10, ipady=5)
        btn_extracao = ttk.Button(buttons_frame, text="Ferramentas de Extração", style='Large.TButton', width=40, bootstyle="primary")
        btn_extracao.config(command=lambda: controller.pulse_and_navigate(btn_extracao, "ExtractionToolsFrame"))
        btn_extracao.pack(pady=10, ipady=5)
        self.right_container = ttk.Frame(self)
        self.right_container.grid(row=1, column=1, sticky='nsew')
        self.watermark_label = ttk.Label(self.right_container)
        self.text_label = ttk.Label(self.right_container)
        self.update_watermarks()

    def update_watermarks(self):
        theme_type = self.controller.app_style.theme.type
        logo_file = 'logo_light.png' if theme_type == 'dark' else 'logo_dark.png'
        text_file = 'logo_text_light.png' if theme_type == 'dark' else 'logo_text_dark.png'
        try:
            watermark_pil_image = Image.open(self.controller.resource_path(logo_file)).resize((450, 450), Image.LANCZOS).convert("RGBA")
            enhancer = ImageEnhance.Brightness(watermark_pil_image)
            watermark_pil_image = enhancer.enhance(0.4)
            alpha = watermark_pil_image.getchannel('A')
            alpha = alpha.point(lambda p: p // 6)
            watermark_pil_image.putalpha(alpha)
            self.watermark_image = ImageTk.PhotoImage(watermark_pil_image)
            self.watermark_label.config(image=self.watermark_image)
            self.watermark_label.pack(expand=True)
            text_pil_image = Image.open(self.controller.resource_path(text_file)).convert("RGBA")
            alpha_text = text_pil_image.getchannel('A')
            alpha_text = alpha_text.point(lambda p: p // 6)
            text_pil_image.putalpha(alpha_text)
            self.text_image = ImageTk.PhotoImage(text_pil_image)
            self.text_label.config(image=self.text_image)
            self.text_label.place(relx=0.5, rely=0.75, anchor='center')
        except Exception as e:
            logging.warning(f"Não foi possível carregar as logos para a marca d'água do menu: {e}")

# --- TELA DE SUB-MENU DE FERRAMENTAS DE EXTRAÇÃO ---
class ExtractionToolsFrame(ttk.Frame):
    def __init__(self, parent, controller, *args): # Adicionado *args
        super().__init__(parent)
        self.controller = controller
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)
        main_container = ttk.Frame(self)
        main_container.pack(expand=True)
        ttk.Label(main_container, text="Ferramentas de Extração", font=("Helvetica", 24, "bold")).pack(pady=(0, 40))
        buttons_frame = ttk.Frame(main_container)
        buttons_frame.pack()
        style = self.controller.app_style
        style.configure('Large.TButton', font=('Helvetica', 12))
        btn_nfe = ttk.Button(buttons_frame, text="Extrator de NFe (XML)", style='Large.TButton', width=40)
        btn_nfe.config(command=lambda: controller.pulse_and_navigate(btn_nfe, "NFeToolFrame"))
        btn_nfe.pack(pady=10, ipady=5)
        btn_futura = ttk.Button(buttons_frame, text="(Futura Ferramenta de Extração)", style='Large.TButton', width=40, state="disabled")
        btn_futura.pack(pady=10, ipady=5)
        btn_voltar = ttk.Button(self, text="< Voltar ao Menu Principal", command=lambda: controller.show_frame("HomeFrame"))
        btn_voltar.place(relx=0.01, rely=0.98, anchor='sw')

# --- TELA DA FERRAMENTA DE NFE ---
class NFeToolFrame(ttk.Frame):
    def __init__(self, parent, controller, *args): # Adicionado *args
        super().__init__(parent)
        self.controller = controller
        self.dados_extraidos_em_memoria = []
        self.contagem_processos = {}
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=0); self.rowconfigure(1, weight=0); self.rowconfigure(2, weight=0)
        self.rowconfigure(3, weight=1); self.rowconfigure(4, weight=0); self.rowconfigure(5, weight=0); self.rowconfigure(6, weight=0)
        self.create_widgets()

    def create_widgets(self):
        self.watermark_label_tool = tk.Label(self, bd=0)
        self.update_watermarks()
        top_frame = ttk.Frame(self)
        top_frame.grid(row=0, column=0, pady=(10,5), sticky='ew')
        top_frame.columnconfigure(1, weight=1)
        btn_voltar = ttk.Button(top_frame, text="< Voltar", command=lambda: self.controller.pulse_and_navigate(btn_voltar, "ExtractionToolsFrame"))
        btn_voltar.grid(row=0, column=0, padx=10, sticky='w')
        title_label = ttk.Label(top_frame, text="Extrator de NFe", font=("Helvetica", 18, "bold"))
        title_label.grid(row=0, column=1)
        path_frame = ttk.Frame(self, padding="10")
        path_frame.grid(row=1, column=0, sticky='ew')
        path_frame.columnconfigure(1, weight=1)
        ttk.Label(path_frame, text="Pasta XML:").grid(row=0, column=0, padx=(0,5), sticky='w')
        self.path_entry = ttk.Entry(path_frame)
        self.path_entry.grid(row=0, column=1, padx=5, sticky='ew')
        ttk.Button(path_frame, text="Selecionar", command=self.selecionar_pasta_origem).grid(row=0, column=2, padx=(5,0))
        self.progress_bar = ttk.Progressbar(self, mode='determinate')
        self.progress_bar.grid(row=2, column=0, sticky='ew', padx=10, pady=5)
        ttk.Frame(self).grid(row=3, column=0) 
        self.status_var = tk.StringVar(value="Pronto.")
        status_frame = ttk.Frame(self, padding=(10, 5))
        status_frame.grid(row=4, column=0, sticky='ew')
        status_frame.columnconfigure(0, weight=1)
        status_label = ttk.Label(status_frame, textvariable=self.status_var, font=("Helvetica", 8))
        status_label.grid(row=0, column=0, sticky='w')
        controls_frame = ttk.Frame(self, padding="10")
        controls_frame.grid(row=5, column=0, pady=5, sticky='ew')
        controls_frame.columnconfigure(0, weight=1)
        controls_frame.columnconfigure(1, weight=1)
        btn_extracao = ttk.Button(controls_frame, text="Importar XML", command=self.extrair_dados_e_salvar_em_memoria, width=30)
        btn_extracao.grid(row=0, column=0, columnspan=2, pady=5, sticky='ew', padx=5)
        self.btn_salvar_basico = ttk.Button(controls_frame, text="Salvar Planilha Básica", command=self.salvar_dados_basicos, state='disabled')
        self.btn_salvar_basico.grid(row=1, column=0, pady=5, padx=(5, 2), sticky='ew')
        self.btn_salvar_completo = ttk.Button(controls_frame, text="Salvar Planilha Completa", command=self.salvar_planilha_completa, state='disabled')
        self.btn_salvar_completo.grid(row=1, column=1, pady=5, padx=(2, 5), sticky='ew')
        btn_limpar = ttk.Button(controls_frame, text="Limpar", command=self.limpar_dados_e_interface)
        btn_limpar.grid(row=2, column=0, pady=5, padx=(5, 2), sticky='ew')
        signature_frame = ttk.Frame(self, padding=(10, 5))
        signature_frame.grid(row=6, column=0, sticky='ew')
        signature_frame.columnconfigure(0, weight=1)
        ttk.Label(signature_frame, text="Desenvolvido por Bruno Silva - Analista Contábil", font=("Helvetica", 8)).pack()
    def update_watermarks(self):
        theme_type = self.controller.app_style.theme.type
        logo_file = 'logo_light.png' if theme_type == 'dark' else 'logo_dark.png'
        try:
            watermark_pil_image = Image.open(self.controller.resource_path(logo_file)).resize((350, 350), Image.LANCZOS).convert("RGBA")
            enhancer = ImageEnhance.Brightness(watermark_pil_image)
            watermark_pil_image = enhancer.enhance(0.4)
            alpha = watermark_pil_image.getchannel('A')
            alpha = alpha.point(lambda p: p // 6)
            watermark_pil_image.putalpha(alpha)
            self.watermark_image_tool = ImageTk.PhotoImage(watermark_pil_image)
            self.watermark_label_tool.config(image=self.watermark_image_tool)
            self.watermark_label_tool.place(relx=0.5, rely=0.45, anchor='center')
        except Exception as e:
            logging.warning(f"Não foi possível carregar a logo para a marca d'água: {e}")
    def update_status(self, message):
        self.status_var.set(message)
        logging.info(message)
        self.controller.update_idletasks()
    def extrair_dados_e_salvar_em_memoria(self):
        try:
            self.dados_extraidos_em_memoria = []
            self.progress_bar['value'] = 0
            self.update_status("Iniciando extração...")
            pasta_xml = self.path_entry.get()
            if not pasta_xml or not os.path.isdir(pasta_xml):
                self.update_status("Erro: Pasta de origem inválida.")
                messagebox.showerror("Erro", "O caminho especificado é inválido. Por favor, selecione uma pasta válida.")
                return
            self.btn_salvar_basico['state'] = 'disabled'
            self.btn_salvar_completo['state'] = 'disabled'
            arquivos_xml = [os.path.join(dp, f) for dp, _, fn in os.walk(pasta_xml) for f in fn if f.lower().endswith('.xml')]
            total_arquivos = len(arquivos_xml)
            self.progress_bar['maximum'] = total_arquivos
            self.update_status(f"Encontrados {total_arquivos} arquivos XML.")
            if total_arquivos == 0:
                messagebox.showinfo("Aviso", "Nenhum arquivo XML encontrado na pasta selecionada.")
                self.update_status("Pronto.")
                return
            arquivos_com_erro = []
            for i, arquivo_path in enumerate(arquivos_xml):
                self.update_status(f"Processando {i + 1}/{total_arquivos}: {os.path.basename(arquivo_path)}")
                self.progress_bar['value'] = i + 1
                dados = core_logic.extrair_dados_nf(arquivo_path)
                if dados: self.dados_extraidos_em_memoria.append(dados)
                else: arquivos_com_erro.append(os.path.basename(arquivo_path))
            if arquivos_com_erro:
                msg = "Falha ao processar os seguintes arquivos:\n" + "\n".join(arquivos_com_erro)
                messagebox.showwarning("Aviso", msg)
            self.update_status(f"Extração concluída. {len(self.dados_extraidos_em_memoria)} arquivos processados.")
            messagebox.showinfo("Concluído", f"Extração concluída!\n{len(self.dados_extraidos_em_memoria)} arquivos processados com sucesso.")
            self.btn_salvar_basico['state'] = 'normal'
            self.btn_salvar_completo['state'] = 'normal'
        except Exception as e:
            logging.error("Ocorreu um erro inesperado durante a extração.", exc_info=True)
            messagebox.showerror("Erro Crítico", f"Ocorreu um erro inesperado:\n{e}\n\nVerifique o arquivo debug.log para detalhes.")
            self.update_status("Erro crítico durante a extração.")
            self.btn_salvar_basico['state'] = 'disabled'
            self.btn_salvar_completo['state'] = 'disabled'
    def salvar_dados_basicos(self):
        if not self.dados_extraidos_em_memoria: messagebox.showerror("Erro", "Nenhum dado extraído."); return
        filename = self.controller.output_filename_pattern.format(data=datetime.now().strftime('%Y-%m-%d'))
        initial_dir = self.controller.default_output_path or os.path.expanduser("~")
        caminho_excel_final = filedialog.asksaveasfilename(
            initialdir=initial_dir,
            initialfile=filename,
            defaultextension=".xlsx", 
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if not caminho_excel_final: return
        self.update_status("Gerando Excel Básico...")
        wb = Workbook()
        saida_ws, entrada_ws = wb.create_sheet("Saída"), wb.create_sheet("Entrada")
        del wb['Sheet']
        cabecalhos_entrada = ['Arquivo', 'Número da NF', 'CFOP', 'Cliente', 'Data de Emissão', 'Valor Total dos Produtos', 'Valor Total', 'Valor II', 'Valor ICMS', 'Valor IPI', 'Valor PIS', 'Valor COFINS', 'Valor AFRMM', 'Outras Despesas', 'Valor Serviço Trading', 'Nome do Processo', 'Status da NF']
        cabecalhos_saida = ['Arquivo', 'Número da NF', 'CFOP', 'Cliente', 'Data de Emissão', 'Valor Total dos Produtos', 'Valor Total', 'Valor ICMS', 'Valor IPI', 'Valor PIS', 'Valor COFINS', 'Valor Serviço Trading', 'Nome do Processo', 'Status da NF']
        core_logic.setup_headers(entrada_ws, cabecalhos_entrada)
        core_logic.setup_headers(saida_ws, cabecalhos_saida)
        row_entrada, row_saida = 2, 2
        for dados in self.dados_extraidos_em_memoria:
            if dados['tipo_nota'] == "Saída":
                core_logic.write_data_to_excel(saida_ws, row_saida, dados, cabecalhos_saida); row_saida += 1
            else:
                core_logic.write_data_to_excel(entrada_ws, row_entrada, dados, cabecalhos_entrada); row_entrada += 1
        core_logic.add_totals_row(entrada_ws, cabecalhos_entrada)
        core_logic.add_totals_row(saida_ws, cabecalhos_saida)
        self.salvar_arquivo(wb, caminho_excel_final)
    def salvar_planilha_completa(self):
        if not self.dados_extraidos_em_memoria: messagebox.showerror("Erro", "Nenhum dado extraído."); return
        processos_saida_encontrados = defaultdict(list)
        for dados in self.dados_extraidos_em_memoria:
            if dados['tipo_nota'] == 'Saída' and dados['nome_processo'] != 'N/A':
                processos_saida_encontrados[dados['nome_processo']].append(dados['numero_nf'])
        processos_para_contar = {proc: len(nfs) for proc, nfs in processos_saida_encontrados.items() if len(nfs) > 1}
        for processo, encontrados in processos_para_contar.items():
            resposta = self.obter_contagem_esperada(processo, encontrados)
            if resposta is None: return
            self.contagem_processos[processo] = resposta
        filename = self.controller.output_filename_pattern.format(data=datetime.now().strftime('%Y-%m-%d'))
        initial_dir = self.controller.default_output_path or os.path.expanduser("~")
        caminho_excel_final = filedialog.asksaveasfilename(
            initialdir=initial_dir,
            initialfile=filename,
            defaultextension=".xlsx", 
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if not caminho_excel_final: return
        self.update_status("Gerando Planilha Completa...")
        wb = Workbook()
        saida_ws, entrada_ws = wb.create_sheet("Saída"), wb.create_sheet("Entrada")
        servico_ws, pendentes_ws = wb.create_sheet("Notas de Serviço Autorizadas"), wb.create_sheet("Notas de Serviço - Pendentes")
        del wb['Sheet']
        cabecalhos_entrada = ['Arquivo', 'Número da NF', 'CFOP', 'Cliente', 'Data de Emissão', 'Valor Total dos Produtos', 'Valor Total', 'Valor II', 'Valor ICMS', 'Valor IPI', 'Valor PIS', 'Valor COFINS', 'Valor AFRMM', 'Outras Despesas', 'Valor Serviço Trading', 'Nome do Processo', 'Status da NF']
        cabecalhos_saida = ['Arquivo', 'Número da NF', 'CFOP', 'Cliente', 'Data de Emissão', 'Valor Total dos Produtos', 'Valor Total', 'Valor ICMS', 'Valor IPI', 'Valor PIS', 'Valor COFINS', 'Valor Serviço Trading', 'Nome do Processo', 'Status da NF']
        cabecalhos_servico = ['Número da NF', 'Cliente', 'Data de Emissão', 'Nome do Processo', 'Valor Serviço Trading']
        core_logic.setup_headers(entrada_ws, cabecalhos_entrada); core_logic.setup_headers(saida_ws, cabecalhos_saida)
        core_logic.setup_headers(servico_ws, cabecalhos_servico); core_logic.setup_headers(pendentes_ws, cabecalhos_servico)
        row_entrada, row_saida, row_servico, row_pendente = 2, 2, 2, 2
        for dados in self.dados_extraidos_em_memoria:
            if dados['tipo_nota'] == "Saída":
                core_logic.write_data_to_excel(saida_ws, row_saida, dados, cabecalhos_saida); row_saida += 1
            else:
                core_logic.write_data_to_excel(entrada_ws, row_entrada, dados, cabecalhos_entrada); row_entrada += 1
            if dados['valor_servico_trading'] != 0.0 and dados['tipo_nota'] == 'Entrada':
                linha_servico_trading = [dados['numero_nf'], dados['nome_cliente'], dados['data_emissao'], dados['nome_processo'], dados['valor_servico_trading']]
                processo_da_nota = dados['nome_processo']
                nfs_encontradas = len(processos_saida_encontrados.get(processo_da_nota, []))
                contagem_esperada = self.contagem_processos.get(processo_da_nota)
                is_authorized = False
                if nfs_encontradas == 1 and contagem_esperada is None: is_authorized = True
                elif contagem_esperada is not None and nfs_encontradas == contagem_esperada: is_authorized = True
                if is_authorized:
                    for col, value in enumerate(linha_servico_trading, 1): servico_ws.cell(row=row_servico, column=col, value=value)
                    row_servico += 1
                else:
                    for col, value in enumerate(linha_servico_trading, 1): pendentes_ws.cell(row=row_pendente, column=col, value=value)
                    row_pendente += 1
        core_logic.add_totals_row(entrada_ws, cabecalhos_entrada)
        core_logic.add_totals_row(saida_ws, cabecalhos_saida)
        core_logic.add_totals_row(servico_ws, cabecalhos_servico)
        core_logic.add_totals_row(pendentes_ws, cabecalhos_servico)
        self.salvar_arquivo(wb, caminho_excel_final)
    def obter_contagem_esperada(self, processo, encontrados):
        dialog = ttk.Toplevel(self.controller)
        dialog.title("Verificação de Processo")
        dialog.transient(self.controller); dialog.grab_set(); dialog.geometry("400x220")
        pos_x = self.controller.winfo_x() + (self.controller.winfo_width() - 400) // 2
        pos_y = self.controller.winfo_y() + (self.controller.winfo_height() - 220) // 2
        dialog.geometry(f"+{pos_x}+{pos_y}")
        main_frame = ttk.Frame(dialog, padding=20); main_frame.pack(expand=True, fill="both")
        ttk.Label(main_frame, text=f"Processo: {processo}", font=("Helvetica", 10, 'bold')).pack(pady=(0, 5))
        ttk.Label(main_frame, text=f"Foram encontradas {encontrados} nota(s) de saída para este processo.").pack(pady=(0, 15))
        input_frame = ttk.Frame(main_frame); input_frame.pack(pady=5)
        ttk.Label(input_frame, text="Qual a quantidade total esperada?").pack(side=tk.LEFT, padx=(0, 10))
        entry = ttk.Entry(input_frame, width=8); entry.pack(side=tk.LEFT); entry.focus_set()
        entry.bind("<Return>", lambda e: on_ok())
        resultado = tk.IntVar(value=-1)
        def on_ok():
            try:
                val = int(entry.get())
                if val > 0:
                    resultado.set(val); dialog.destroy()
                else:
                    messagebox.showwarning("Valor Inválido", "Por favor, insira um número inteiro positivo.", parent=dialog)
            except (ValueError, TypeError):
                messagebox.showwarning("Valor Inválido", "Por favor, insira um número válido.", parent=dialog)
        def on_cancel():
            dialog.destroy()
        dialog.bind("<Escape>", lambda e: on_cancel())
        button_frame = ttk.Frame(main_frame); button_frame.pack(pady=20)
        ok_button = ttk.Button(button_frame, text="Confirmar", command=on_ok, bootstyle="primary")
        ok_button.pack(side=tk.LEFT, padx=10)
        cancel_button = ttk.Button(button_frame, text="Cancelar", command=on_cancel)
        cancel_button.pack(side=tk.LEFT)
        self.controller.wait_window(dialog)
        return resultado.get() if resultado.get() != -1 else None
    def salvar_arquivo(self, wb, caminho):
        try:
            wb.save(caminho)
            self.update_status(f"Arquivo salvo com sucesso!")
            if self.controller.ask_to_open_excel:
                if messagebox.askyesno("Abrir Arquivo", "Planilha salva com sucesso!\nDeseja abri-la agora?"):
                    os.startfile(caminho)
        except Exception as e:
            self.update_status(f"Erro ao salvar o arquivo Excel.")
            logging.error(f"Erro ao salvar o arquivo Excel: {e}", exc_info=True)
            messagebox.showerror("Erro", f"Erro ao salvar o arquivo Excel: {e}")
    def limpar_dados_e_interface(self):
        if messagebox.askyesno("Confirmação", "Deseja realmente limpar todos os dados e reiniciar?"):
            self.dados_extraidos_em_memoria.clear()
            self.path_entry.delete(0, tk.END)
            self.progress_bar.config(value=0)
            self.update_status("Pronto.")
            self.btn_salvar_basico['state'] = 'disabled'
            self.btn_salvar_completo['state'] = 'disabled'
            logging.info("Interface e dados limpos pelo usuário.")
    def selecionar_pasta_origem(self):
        initial_dir = self.controller.default_xml_path or os.path.expanduser("~")
        folder_path = filedialog.askdirectory(title="Selecione a pasta com os arquivos XML", initialdir=initial_dir)
        if folder_path:
            self.path_entry.delete(0, tk.END)
            self.path_entry.insert(0, folder_path)
            self.update_status(f"Pasta selecionada: {folder_path}")