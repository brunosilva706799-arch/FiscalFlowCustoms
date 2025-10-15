# =============================================================================
# --- ARQUIVO: core_logic.py (COM EXTRAÇÃO DE PROCESSO APRIMORADA) ---
# =============================================================================
import os
import re
import logging
from collections import defaultdict
from lxml import etree
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import requests

def _get_namespace(root_xml):
    """Extrai o namespace do elemento raiz do XML."""
    if not hasattr(root_xml, 'tag') or not isinstance(root_xml.tag, str):
        return ''
    namespace_match = re.search(r'\{(.*?)\}', root_xml.tag)
    return namespace_match.group(1) if namespace_match else ''

def _normalize_processo(text):
    """
    Recebe uma string de texto e a limpa para criar um identificador comum.
    Ex: 'PROCESSO EMALI.IMP001/2025' -> 'EMALIIMP0012025'
    """
    if not text or not isinstance(text, str):
        return 'N/A'
    
    clean_text = text.upper().replace("PROCESSO", "").replace(":", "").replace(".", "").replace("/", "").replace("-", "").replace(" ", "")
    return clean_text if clean_text else 'N/A'

# --- [NOVA FUNÇÃO DE DIAGNÓSTICO] ---
def gerar_diagnostico_processos(dados_extraidos, caminho_saida):
    """Gera um arquivo de texto para diagnosticar problemas de vínculo de processo."""
    try:
        with open(caminho_saida, 'w', encoding='utf-8') as f:
            f.write("Relatório de Diagnóstico de Processos\n")
            f.write("="*40 + "\n\n")
            
            dados_ordenados = sorted(dados_extraidos, key=lambda x: (x.get('processo_normalizado', ''), x.get('tipo_nota', '')))

            for dados in dados_ordenados:
                linha = (
                    f"Arquivo: {dados.get('nome_arquivo', 'N/A'):<50} | "
                    f"Tipo: {dados.get('tipo_nota', 'N/A'):<7} | "
                    f"Processo Original: \"{dados.get('nome_processo', 'N/A')}\" | "
                    f"Processo Normalizado: \"{dados.get('processo_normalizado', 'N/A')}\"\n"
                )
                f.write(linha)
        return True, "Diagnóstico gerado com sucesso."
    except Exception as e:
        return False, f"Erro ao gerar diagnóstico: {e}"

def _extrair_dados_autorizada(root_xml, namespace, arquivo_xml):
    def find(element, tag):
        if namespace: return element.find(f'{{{namespace}}}{tag}')
        return element.find(tag)
    def find_all(element, tag):
        if namespace: return element.findall(f'{{{namespace}}}{tag}')
        return element.findall(tag)
    nfe_element = find(root_xml, 'NFe')
    if nfe_element is None: nfe_element = root_xml
    inf_nfe_tag = find(nfe_element, 'infNFe')
    if inf_nfe_tag is None: return None
    ide_tag = find(inf_nfe_tag, 'ide')
    if ide_tag is None: return None
    sistema_emissor = 'Não identificado'
    ver_proc_tag = find(ide_tag, 'verProc')
    if ver_proc_tag is not None and ver_proc_tag.text:
        ver_proc_text = ver_proc_tag.text.strip()
        if re.search(r'[a-zA-Z]', ver_proc_text) and not ver_proc_text.replace('.', '').isdigit():
            sistema_emissor = ver_proc_text
        else:
            inf_resp_tec_tag = find(inf_nfe_tag, 'infRespTec')
            if inf_resp_tec_tag is not None:
                email_tag = find(inf_resp_tec_tag, 'email')
                if email_tag is not None and email_tag.text:
                    match = re.search(r'@([\w.-]+)', email_tag.text)
                    if match:
                        domain = match.group(1); company_name = domain.split('.')[0]
                        if "narwal" in company_name.lower(): sistema_emissor = "Narwal Sistemas"
                        else: sistema_emissor = company_name.capitalize() + " Sistemas"
    numero_nf = find(ide_tag, 'nNF').text if find(ide_tag, 'nNF') is not None else 'N/A'
    try:
        cfop_tag_det = find_all(inf_nfe_tag, 'det')
        if cfop_tag_det:
            prod_tag = find(cfop_tag_det[0], 'prod')
            if prod_tag is not None: cfop_nf = find(prod_tag, 'CFOP').text
        else: cfop_nf = 'N/A'
    except (AttributeError, IndexError): cfop_nf = 'N/A'
    data_emissao_completa = find(ide_tag, 'dhEmi').text if find(ide_tag, 'dhEmi') is not None else 'N/A'
    data_emissao = data_emissao_completa.split('T')[0] if data_emissao_completa != 'N/A' else 'N/A'
    tipo_nota_val = find(ide_tag, 'tpNF').text if find(ide_tag, 'tpNF') is not None else '1'
    tipo_nota = "Saída" if tipo_nota_val == '1' else "Entrada"
    dest_tag = find(inf_nfe_tag, 'dest')
    nome_cliente, razao_social_destinatario, cnpj_cpf_destinatario, uf_destinatario = 'N/A', 'N/A', 'N/A', 'N/A'
    if dest_tag is not None:
        if find(dest_tag, 'xNome') is not None and find(dest_tag, 'xNome').text:
            nome_cliente = find(dest_tag, 'xNome').text; razao_social_destinatario = find(dest_tag, 'xNome').text
        if find(dest_tag, 'CNPJ') is not None and find(dest_tag, 'CNPJ').text:
            cnpj_cpf_destinatario = find(dest_tag, 'CNPJ').text
        elif find(dest_tag, 'CPF') is not None and find(dest_tag, 'CPF').text:
            cnpj_cpf_destinatario = find(dest_tag, 'CPF').text
        ender_dest_tag = find(dest_tag, 'enderDest')
        if ender_dest_tag is not None:
            uf_tag = find(ender_dest_tag, 'UF')
            if uf_tag is not None and uf_tag.text is not None: uf_destinatario = uf_tag.text
    total_tag = find(inf_nfe_tag, 'total'); valor_total_nf, valor_total_produtos = 'N/A', 'N/A'
    impostos_totais = defaultdict(lambda: '0.00'); v_ii, v_outras, v_afrmm = 0.0, 0.0, 0.0
    if total_tag is not None:
        icms_total_tag = find(total_tag, 'ICMSTot')
        if icms_total_tag is not None:
            valor_total_nf = find(icms_total_tag, 'vNF').text if find(icms_total_tag, 'vNF') is not None else 'N/A'
            valor_total_produtos = find(icms_total_tag, 'vProd').text if find(icms_total_tag, 'vProd') is not None else 'N/A'
            for imposto_tag in icms_total_tag:
                tag_nome = imposto_tag.tag.split('}')[-1]
                if tag_nome.startswith('v') and imposto_tag.text: impostos_totais[tag_nome] = imposto_tag.text
            ii_tag = find(icms_total_tag, 'vII')
            if ii_tag is not None and ii_tag.text: v_ii = float(ii_tag.text)
            v_outras_tag = find(icms_total_tag, 'vOutro')
            if v_outras_tag is not None and v_outras_tag.text: v_outras = float(v_outras_tag.text)
    det_tags = find_all(inf_nfe_tag, 'det')
    for det_tag in det_tags:
        prod_tag = find(det_tag, 'prod')
        if prod_tag is not None:
            di_tag = find(prod_tag, 'DI')
            if di_tag is not None:
                vafrmm_tag = find(di_tag, 'vAFRMM')
                if vafrmm_tag is not None and vafrmm_tag.text:
                    try:
                        v_afrmm = float(vafrmm_tag.text)
                        if v_afrmm > 0: break 
                    except (ValueError, TypeError): continue
    inf_adicionais_tag = find(inf_nfe_tag, 'infAdic'); texto_completo_adicional = ""
    if inf_adicionais_tag is not None and find(inf_adicionais_tag, 'infCpl') is not None and find(inf_adicionais_tag, 'infCpl').text:
        texto_completo_adicional = find(inf_adicionais_tag, 'infCpl').text
    if v_afrmm == 0.0 and texto_completo_adicional:
        afrmm_match = re.search(r'AFRMM.*?R\$?\s*([\d.,]+)', texto_completo_adicional, re.IGNORECASE | re.DOTALL)
        if afrmm_match: valor_str = afrmm_match.group(1).replace('.', '').replace(',', '.'); v_afrmm = float(valor_str)
    valor_frete_nacional = 0.0
    if tipo_nota == "Entrada" and texto_completo_adicional:
        frete_match = re.search(r'FRETE\s*NACIONAL.*?[R\$]?\s*([\d.,]+)', texto_completo_adicional, re.IGNORECASE | re.DOTALL)
        if frete_match:
            try:
                valor_str = frete_match.group(1).replace('.', '').replace(',', '.'); valor_frete_nacional = float(valor_str)
            except (ValueError, TypeError) as e:
                logging.warning(f"Erro ao converter frete: {e}"); valor_frete_nacional = 0.0
    nome_processo = 'N/A'
    match_principal = re.search(r'([A-Z]+(?:IMP|EXP)\d+)', texto_completo_adicional.replace("/", "").replace(".",""), re.IGNORECASE)
    if match_principal:
        nome_processo = match_principal.group(1)
    else:
        match_processo = re.search(r'PROCESSO\s*:?\s*([A-Z0-9\./\s-]+)', texto_completo_adicional, re.IGNORECASE)
        if match_processo: nome_processo = match_processo.group(1).strip()
        else:
            for item_tag in find_all(inf_nfe_tag, 'det'):
                inf_ad_prod_tag = find(item_tag, 'infAdProd')
                if inf_ad_prod_tag is not None and inf_ad_prod_tag.text:
                    match_item = re.search(r'(?:PROCESSO|REGISTRO):\s*([A-Z0-9\./\s-]+)', inf_ad_prod_tag.text, re.IGNORECASE)
                    if match_item: nome_processo = match_item.group(1).strip(); break
    processo_normalizado = _normalize_processo(nome_processo)
    valor_servico_trading = 'N/A'
    if texto_completo_adicional:
        match_trading = re.search(r'trading[^0-9]*([\d.,]+)', texto_completo_adicional.lower())
        if match_trading: valor_servico_trading = match_trading.group(1).replace('.', '').replace(',', '.')
    status_nf = "N/A"
    prot_nfe_tag = find(root_xml, 'protNFe')
    if prot_nfe_tag is not None and find(prot_nfe_tag, 'infProt') is not None:
        cstat_tag = find(find(prot_nfe_tag, 'infProt'), 'cStat')
        if cstat_tag is not None:
            if cstat_tag.text == '100': status_nf = "Autorizada"
            elif cstat_tag.text == '101': status_nf = "Cancelada"
    return {
        'nome_arquivo': os.path.basename(arquivo_xml), 'numero_nf': numero_nf, 'cfop_nf': cfop_nf, 'data_emissao': data_emissao,
        'nome_cliente': nome_cliente, 'valor_total_nf': float(valor_total_nf) if valor_total_nf != 'N/A' else 0.0,
        'valor_total_produtos': float(valor_total_produtos) if valor_total_produtos != 'N/A' else 0.0, 'impostos': impostos_totais,
        'valor_servico_trading': float(valor_servico_trading) if valor_servico_trading != 'N/A' else 0.0, 'tipo_nota': tipo_nota,
        'nome_processo': nome_processo, 'vII': v_ii, 'vAFRMM': v_afrmm, 'vOutras': v_outras, 'status_nf': status_nf,
        'sistema_emissor': sistema_emissor, 'uf_destinatario': uf_destinatario, 'cnpj_cpf_destinatario': cnpj_cpf_destinatario,
        'razao_social_destinatario': razao_social_destinatario, 'processo_normalizado': processo_normalizado,
        'valor_frete_nacional': valor_frete_nacional
    }
def _extrair_dados_cancelamento(root_xml, namespace, arquivo_xml):
    def find(element, tag): return element.find(f'{{{namespace}}}{tag}') if namespace else element.find(tag)
    inf_evento = find(root_xml, 'evento');
    if inf_evento is None: return None
    inf_evento = find(inf_evento, 'infEvento');
    if inf_evento is None: return None
    chNFe = find(inf_evento, 'chNFe').text if find(inf_evento, 'chNFe') is not None else 'N/A'
    dhEvento = find(inf_evento, 'dhEvento').text if find(inf_evento, 'dhEvento') is not None else 'N/A'
    data_cancelamento = dhEvento.split('T')[0] if dhEvento != 'N/A' else 'N/A'
    numero_nf_da_chave = chNFe[25:34] if len(chNFe) == 44 else 'N/A'
    return {
        'nome_arquivo': os.path.basename(arquivo_xml), 'numero_nf': numero_nf_da_chave, 'cfop_nf': 'N/A', 'data_emissao': data_cancelamento,
        'nome_cliente': 'N/A (Cancelada)', 'valor_total_nf': 0.0, 'valor_total_produtos': 0.0, 'impostos': defaultdict(lambda: '0.00'),
        'valor_servico_trading': 0.0, 'tipo_nota': 'N/A', 'nome_processo': 'N/A', 'vII': 0.0, 'vAFRMM': 0.0, 'vOutras': 0.0,
        'status_nf': "Cancelada", 'sistema_emissor': 'N/A', 'uf_destinatario': 'N/A', 'cnpj_cpf_destinatario': 'N/A',
        'razao_social_destinatario': 'N/A (Cancelada)', 'processo_normalizado': 'N/A', 'valor_frete_nacional': 0.0
    }
def extrair_dados_nf(arquivo_xml):
    try:
        tree = etree.parse(arquivo_xml); root_xml = tree.getroot(); namespace = _get_namespace(root_xml)
        root_tag = root_xml.tag.split('}')[-1] if '}' in root_xml.tag else root_xml.tag
        if root_tag == 'procEventoNFe':
            inf_evento = root_xml.find(f'.//{{{namespace}}}infEvento') if namespace else root_xml.find('.//infEvento')
            if inf_evento is not None:
                tpEvento_tag = inf_evento.find(f'{{{namespace}}}tpEvento') if namespace else inf_evento.find('tpEvento')
                if tpEvento_tag is not None and tpEvento_tag.text == '110111':
                    return _extrair_dados_cancelamento(root_xml, namespace, arquivo_xml)
        elif root_tag in ['nfeProc', 'NFe']:
            return _extrair_dados_autorizada(root_xml, namespace, arquivo_xml)
        logging.info(f"Arquivo ignorado: {os.path.basename(arquivo_xml)}")
        return None
    except etree.XMLSyntaxError: logging.error(f"Erro de sintaxe XML: {os.path.basename(arquivo_xml)}."); return None
    except Exception as e: logging.error(f"Erro inesperado: {os.path.basename(arquivo_xml)}: {e}", exc_info=True); return None
def setup_headers(ws, headers):
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"); alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title); cell.fill, cell.alignment, cell.border = fill, alignment, border
def write_data_to_excel(ws, row_index, data, headers):
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')); currency_format = 'R$ #,##0.00'
    for col, header in enumerate(headers, 1):
        value_map = {
            'Arquivo': data.get('nome_arquivo'), 'Número da NF': data.get('numero_nf'), 'CFOP': data.get('cfop_nf'), 'Cliente': data.get('nome_cliente'),
            'Data de Emissão': data.get('data_emissao'), 'Nome do Processo': data.get('nome_processo'), 'Valor Total dos Produtos': data.get('valor_total_produtos', 0.0),
            'Valor Total': data.get('valor_total_nf', 0.0), 'Valor ICMS': float(data.get('impostos', {}).get('vICMS', '0.00')),
            'Valor IPI': float(data.get('impostos', {}).get('vIPI', '0.00')), 'Valor PIS': float(data.get('impostos', {}).get('vPIS', '0.00')),
            'Valor COFINS': float(data.get('impostos', {}).get('vCOFINS', '0.00')), 'Valor Serviço Trading': data.get('valor_servico_trading', 0.0),
            'Valor II': data.get('vII', 0.0), 'Valor AFRMM': data.get('vAFRMM', 0.0), 'Outras Despesas': data.get('vOutras', 0.0),
            'Status da NF': data.get('status_nf', 'N/A'), 'Sistema Emissor': data.get('sistema_emissor', 'N/A'), 'UF Destino': data.get('uf_destinatario'),
            'CNPJ/CPF Dest.': data.get('cnpj_cpf_destinatario'), 'Razão Social Dest.': data.get('razao_social_destinatario'),
            'Frete Nacional': data.get('valor_frete_nacional', 0.0)
        }
        value = value_map.get(header, '')
        cell = ws.cell(row=row_index, column=col, value=value); cell.border = border
        if isinstance(value, (int, float)) and (header.startswith('Valor') or header in ['Outras Despesas', 'Valor AFRMM', 'Valor Serviço Trading', 'Frete Nacional']):
            cell.number_format = currency_format
def add_totals_row(ws, headers):
    last_row = ws.max_row
    if last_row < 2: return
    totals_row = last_row + 1; total_font = Font(bold=True); total_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    first_cell = ws.cell(row=totals_row, column=1, value="TOTAIS"); first_cell.font, first_cell.fill = total_font, total_fill
    for col_idx, header in enumerate(headers, 1):
        if header.startswith('Valor') or header in ['Outras Despesas', 'Valor AFRMM', 'Valor Serviço Trading', 'Frete Nacional']:
            col_letter = get_column_letter(col_idx); formula = f"=SUM({col_letter}2:{col_letter}{last_row})"
            cell = ws.cell(row=totals_row, column=col_idx, value=formula); cell.font, cell.fill, cell.number_format = total_font, total_fill, 'R$ #,##0.00'
def check_for_updates(current_version, repo_owner, repo_name):
    api_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/releases/latest"
    try:
        response = requests.get(api_url, timeout=5); response.raise_for_status()
        latest_release = response.json(); latest_version_str = latest_release['tag_name'].lstrip('v')
        current_v, latest_v = tuple(map(int, (current_version.split('.')))), tuple(map(int, (latest_version_str.split('.'))))
        if latest_v > current_v:
            assets = latest_release.get('assets', [])
            if assets:
                download_url = assets[0].get('browser_download_url')
                return {"update_available": True, "latest_version": latest_version_str, "download_url": download_url, "release_notes": latest_release.get("body", "Sem notas.")}
    except Exception as e:
        logging.error(f"Erro ao verificar updates: {e}")
    return {"update_available": False}
def download_update(url, progress_callback, completion_callback, error_callback):
    try:
        local_filename = url.split('/')[-1]
        download_path = os.path.join(os.path.expanduser('~'), 'Downloads', local_filename)
        with requests.get(url, stream=True) as r:
            r.raise_for_status()
            total_size_in_bytes = int(r.headers.get('content-length', 0)); downloaded_bytes = 0
            with open(download_path, 'wb') as f:
                for chunk in r.iter_content(chunk_size=8192):
                    f.write(chunk); downloaded_bytes += len(chunk)
                    progress_callback(downloaded_bytes, total_size_in_bytes)
        completion_callback(download_path)
    except Exception as e:
        logging.error(f"Erro no download: {e}", exc_info=True); error_callback(str(e))
def calcular_dados_dashboard(dados_extraidos):
    if not dados_extraidos: return {'resumo_geral': {}, 'processos_pendentes': {}}
    resumo = {'total_notas': 0, 'notas_entrada': 0, 'notas_saida': 0, 'notas_canceladas': 0, 'valor_total_entrada': 0.0, 'valor_total_saida': 0.0, 'valor_total_produtos': 0.0, 'impostos': {'Entrada': defaultdict(float), 'Saída': defaultdict(float)}, 'processos_unicos': set()}
    processos_saida_contagem = defaultdict(int)
    for nf in dados_extraidos:
        resumo['total_notas'] += 1
        if 'Cancelada' in nf.get('status_nf', ''):
            resumo['notas_canceladas'] += 1; continue
        valor_da_nota, impostos_nf = nf.get('valor_total_nf', 0.0), nf.get('impostos', {})
        tipo_operacao = nf.get('tipo_nota', 'Saída')
        if tipo_operacao == "Entrada": resumo['notas_entrada'] += 1; resumo['valor_total_entrada'] += valor_da_nota
        else: resumo['notas_saida'] += 1; resumo['valor_total_saida'] += valor_da_nota
        resumo['valor_total_produtos'] += nf.get('valor_total_produtos', 0.0)
        for imp in ['vICMS', 'vIPI', 'vPIS', 'vCOFINS']:
            resumo['impostos'][tipo_operacao][imp] += float(impostos_nf.get(imp, '0.0'))
        processo_norm = nf.get('processo_normalizado', 'N/A')
        if processo_norm and processo_norm != 'N/A':
            resumo['processos_unicos'].add(processo_norm)
            if tipo_operacao == "Saída": processos_saida_contagem[processo_norm] += 1
    processos_pendentes = {p: c for p, c in processos_saida_contagem.items() if c > 1}
    return {'resumo_geral': resumo, 'processos_pendentes': processos_pendentes}