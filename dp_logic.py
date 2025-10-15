# =============================================================================
# --- ARQUIVO: dp_logic.py ---
# (Atualizado com o formato de data YYYYMM no arquivo TXT)
# =============================================================================

from firebase_admin import firestore
import openpyxl
import logging
import re

# --- LÓGICA PARA EMPRESAS (COMPANIES) ---

def get_all_companies():
    """Busca todas as empresas cadastradas no Firestore."""
    db = firestore.client()
    try:
        companies_list = []
        companies_ref = db.collection('companies').order_by('name').stream()
        for company in companies_ref:
            company_data = company.to_dict()
            companies_list.append({
                'id': company.id,
                'code': company_data.get('code'),
                'name': company_data.get('name'),
                'cnpj': company_data.get('cnpj')
            })
        return companies_list, None
    except Exception as e:
        error_message = f"ERRO AO BUSCAR EMPRESAS NO FIREBASE:\n\n{e}"
        print(error_message)
        return [], error_message

def add_company(code, name, cnpj):
    """Adiciona uma nova empresa ao Firestore."""
    db = firestore.client()
    try:
        existing_company = db.collection('companies').where('code', '==', code).limit(1).stream()
        if len(list(existing_company)) > 0:
            return "Erro: Já existe uma empresa cadastrada com este código."
        
        db.collection('companies').add({
            'code': code,
            'name': name,
            'cnpj': cnpj
        })
        return "Empresa adicionada com sucesso."
    except Exception as e:
        return f"Ocorreu um erro ao adicionar a empresa: {e}"

def update_company(company_id, code, name, cnpj):
    """Atualiza os dados de uma empresa existente."""
    db = firestore.client()
    try:
        existing_company_query = db.collection('companies').where('code', '==', code).limit(1).stream()
        for company in existing_company_query:
            if company.id != company_id:
                return "Erro: O novo código já está em uso por outra empresa."

        db.collection('companies').document(company_id).update({
            'code': code,
            'name': name,
            'cnpj': cnpj
        })
        return "Empresa atualizada com sucesso."
    except Exception as e:
        return f"Ocorreu um erro ao atualizar a empresa: {e}"

def delete_company(company_id):
    """Remove uma empresa do Firestore."""
    db = firestore.client()
    try:
        db.collection('companies').document(company_id).delete()
        return "Empresa removida com sucesso."
    except Exception as e:
        return f"Ocorreu um erro ao remover a empresa: {e}"

# --- LÓGICA PARA COLABORADORES (EMPLOYEES) ---

def get_employees_for_company(company_id):
    """Busca todos os colaboradores de uma empresa específica."""
    db = firestore.client()
    try:
        employees_list = []
        employees_ref = db.collection('employees').where('company_id', '==', company_id).order_by('full_name').stream()
        for emp in employees_ref:
            emp_data = emp.to_dict()
            employees_list.append({
                'id': emp.id,
                'employee_code': emp_data.get('employee_code'),
                'full_name': emp_data.get('full_name'),
                'salary': emp_data.get('salary', 0.0)
            })
        return employees_list, None
    except Exception as e:
        error_message = f"ERRO AO BUSCAR COLABORADORES NO FIREBASE:\n\n{e}"
        print(error_message)
        return [], error_message

def add_employee(company_id, employee_code, full_name, salary):
    """Adiciona um novo colaborador a uma empresa."""
    db = firestore.client()
    try:
        query = db.collection('employees').where('company_id', '==', company_id).where('employee_code', '==', employee_code).limit(1).stream()
        if len(list(query)) > 0:
            return f"Erro: O código '{employee_code}' já está em uso nesta empresa."
        
        db.collection('employees').add({
            'company_id': company_id,
            'employee_code': employee_code,
            'full_name': full_name,
            'salary': salary
        })
        return "Colaborador adicionado com sucesso."
    except Exception as e:
        return f"Ocorreu um erro ao adicionar o colaborador: {e}"

def update_employee(employee_id, employee_code, full_name, salary, company_id):
    """Atualiza os dados de um colaborador."""
    db = firestore.client()
    try:
        query = db.collection('employees').where('company_id', '==', company_id).where('employee_code', '==', employee_code).limit(1).stream()
        for emp in query:
            if emp.id != employee_id:
                return f"Erro: O código '{employee_code}' já está em uso por outro colaborador nesta empresa."

        db.collection('employees').document(employee_id).update({
            'employee_code': employee_code,
            'full_name': full_name,
            'salary': salary
        })
        return "Colaborador atualizado com sucesso."
    except Exception as e:
        return f"Ocorreu um erro ao atualizar o colaborador: {e}"

def delete_multiple_employees(employee_ids):
    """Remove uma lista de colaboradores do Firestore usando uma operação em lote."""
    db = firestore.client()
    try:
        if not employee_ids:
            return (False, "Nenhum colaborador selecionado para remoção.")

        batch = db.batch()
        employees_ref = db.collection('employees')
        
        for emp_id in employee_ids:
            doc_ref = employees_ref.document(emp_id)
            batch.delete(doc_ref)
        
        batch.commit()
        
        return (True, f"{len(employee_ids)} colaborador(es) removido(s) com sucesso.")
    except Exception as e:
        logging.error(f"Erro ao remover múltiplos colaboradores: {e}", exc_info=True)
        return (False, f"Ocorreu um erro ao remover os colaboradores:\n\n{e}")

def read_employees_from_file(company_id, file_path):
    """
    Lê uma planilha .xlsx de forma inteligente, processa os dados e os retorna
    para uma tela de pré-visualização, sem salvar no banco.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        header_row_num = -1
        code_col_idx, name_col_idx, salary_col_idx = -1, -1, -1

        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            if row is None: continue
            
            has_code, has_name = False, False
            for j, cell_value in enumerate(row):
                if cell_value and isinstance(cell_value, str):
                    cv_lower = cell_value.lower()
                    if "código" in cv_lower:
                        has_code = True; code_col_idx = j
                    if "nome" in cv_lower:
                        has_name = True; name_col_idx = j
                    if "salário" in cv_lower:
                        salary_col_idx = j
            
            if has_code and has_name:
                header_row_num = i
                break
        
        if header_row_num == -1:
            return (False, "Erro: Não foi possível encontrar a linha de cabeçalho com 'Código' e 'Nome' na planilha.")

        existing_employees, error = get_employees_for_company(company_id)
        if error:
            return (False, "Erro ao buscar funcionários existentes no banco de dados.")
        existing_codes = {emp['employee_code'] for emp in existing_employees}

        processed_employees = []
        for i, row in enumerate(sheet.iter_rows(min_row=header_row_num + 1, values_only=True), start=header_row_num + 1):
            if row is None: continue
            if len(row) <= max(code_col_idx, name_col_idx): continue

            raw_code = row[code_col_idx]
            raw_name = row[name_col_idx]
            raw_salary = row[salary_col_idx] if salary_col_idx != -1 and len(row) > salary_col_idx else 0.0

            employee_code = None
            if raw_code is not None:
                try:
                    employee_code = str(int(float(raw_code)))
                except (ValueError, TypeError):
                    employee_code = str(raw_code).strip()
            
            if not employee_code or not employee_code.isdigit():
                continue
            
            full_name = str(raw_name).strip() if raw_name is not None else ""
            
            salary = 0.0
            if raw_salary is not None:
                try:
                    salary = float(str(raw_salary).replace(",", "."))
                except (ValueError, TypeError):
                    salary = 0.0

            status = "Ignorado (já existe)" if employee_code in existing_codes else "Novo"
            
            processed_employees.append({
                'employee_code': employee_code,
                'full_name': full_name,
                'salary': salary,
                'status': status
            })
            
        return (True, processed_employees)
    
    except PermissionError:
        return (False, "Erro de Permissão: A planilha não pode ser lida.\n\nPor favor, feche o arquivo no Excel e tente novamente.")
    except Exception as e:
        return (False, f"Ocorreu um erro ao ler a planilha:\n\n{e}")

def save_imported_employees(company_id, employee_list):
    """Recebe uma lista de colaboradores e salva apenas os 'Novos' no Firestore."""
    db = firestore.client()
    try:
        added_count = 0
        batch = db.batch()
        employees_ref = db.collection('employees')

        for emp in employee_list:
            if emp['status'] == "Novo":
                new_doc_ref = employees_ref.document()
                batch.set(new_doc_ref, {
                    'company_id': company_id,
                    'employee_code': emp['employee_code'],
                    'full_name': emp['full_name'],
                    'salary': emp.get('salary', 0.0)
                })
                added_count += 1
        
        batch.commit()
        
        return (True, f"{added_count} novo(s) colaborador(es) foram importados com sucesso.")
    except Exception as e:
        return (False, f"Ocorreu um erro ao salvar os dados no banco de dados:\n\n{e}")


# --- LÓGICA PARA RUBRICAS (PAYROLL_CODES) ---
def get_all_payroll_codes():
    """Busca todas as rubricas cadastradas no Firestore."""
    db = firestore.client()
    try:
        codes_list = []
        codes_ref = db.collection('payroll_codes').order_by('code').stream()
        for code in codes_ref:
            code_data = code.to_dict()
            codes_list.append({
                'id': code.id,
                'code': code_data.get('code'),
                'name': code_data.get('name'),
                'value_type': code_data.get('value_type'),
                'calculation_base': code_data.get('calculation_base', 'Valor Informado'),
                'calculation_factor': code_data.get('calculation_factor', 1.0)
            })
        return codes_list, None
    except Exception as e:
        return [], f"ERRO AO BUSCAR RUBRICAS NO FIREBASE:\n\n{e}"

def add_payroll_code(code, name, value_type, calculation_base, calculation_factor):
    """Adiciona uma nova rubrica ao Firestore."""
    db = firestore.client()
    try:
        query = db.collection('payroll_codes').where('code', '==', code).limit(1).stream()
        if len(list(query)) > 0:
            return f"Erro: O código de rubrica '{code}' já existe."
        
        db.collection('payroll_codes').add({
            'code': code,
            'name': name,
            'value_type': value_type,
            'calculation_base': calculation_base,
            'calculation_factor': calculation_factor
        })
        return "Rubrica adicionada com sucesso."
    except Exception as e:
        return f"Ocorreu um erro ao adicionar a rubrica: {e}"

def update_payroll_code(code_id, code, name, value_type, calculation_base, calculation_factor):
    """Atualiza os dados de uma rubrica."""
    db = firestore.client()
    try:
        query = db.collection('payroll_codes').where('code', '==', code).limit(1).stream()
        for item in query:
            if item.id != code_id:
                return f"Erro: O código de rubrica '{code}' já está em uso."

        db.collection('payroll_codes').document(code_id).update({
            'code': code,
            'name': name,
            'value_type': value_type,
            'calculation_base': calculation_base,
            'calculation_factor': calculation_factor
        })
        return "Rubrica atualizada com sucesso."
    except Exception as e:
        return f"Ocorreu um erro ao atualizar a rubrica: {e}"

def delete_payroll_code(code_id):
    """Remove uma rubrica."""
    db = firestore.client()
    try:
        db.collection('payroll_codes').document(code_id).delete()
        return "Rubrica removida com sucesso."
    except Exception as e:
        return f"Ocorreu um erro ao remover a rubrica: {e}"

def read_payroll_codes_from_file(file_path):
    """Lê uma planilha .xlsx de forma inteligente para importar rubricas."""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        header_row_num = -1
        code_col_idx, name_col_idx, type_col_idx = -1, -1, -1

        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            if row is None: continue
            
            has_code, has_name, has_type = False, False, False
            for j, cell_value in enumerate(row):
                str_value = str(cell_value).strip().lower() if cell_value is not None else ""
                if "cód." in str_value:
                    has_code = True; code_col_idx = j
                if "descrição" in str_value:
                    has_name = True; name_col_idx = j
                if "unidade" in str_value:
                    has_type = True; type_col_idx = j
            
            if has_code and has_name and has_type:
                header_row_num = i
                break
        
        if header_row_num == -1:
            return (False, "Erro: Não foi possível encontrar os cabeçalhos 'Cód.', 'Descrição' e 'Unidade' na planilha.")

        existing_codes_data, error = get_all_payroll_codes()
        if error:
            return (False, "Erro ao buscar rubricas existentes no banco de dados.")
        existing_codes = {str(c['code']) for c in existing_codes_data}

        processed_codes = []
        for i, row in enumerate(sheet.iter_rows(min_row=header_row_num + 1, values_only=True), start=header_row_num + 1):
            if row is None: continue

            max_idx_needed = max(code_col_idx, name_col_idx, type_col_idx)
            if len(row) <= max_idx_needed:
                continue

            raw_code = row[code_col_idx]
            raw_name = row[name_col_idx]
            raw_type = row[type_col_idx]

            code = None
            if raw_code is not None:
                try:
                    code = str(int(float(raw_code)))
                except (ValueError, TypeError):
                    code = str(raw_code).strip()
            
            if not code or not code.isdigit():
                continue
            
            name = str(raw_name).strip() if raw_name is not None else ""
            value_type = str(raw_type).strip() if raw_type is not None else "Valor"
            
            status = "Ignorado (já existe)" if code in existing_codes else "Novo"
            
            processed_codes.append({
                'code': code,
                'name': name,
                'value_type': value_type,
                'status': status,
                'calculation_base': 'Valor Informado',
                'calculation_factor': 1.0
            })
            
        return (True, processed_codes)

    except PermissionError:
        return (False, "Erro de Permissão: A planilha não pode ser lida.\n\nPor favor, feche o arquivo no Excel e tente novamente.")
    except Exception as e:
        return (False, f"Ocorreu um erro ao ler a planilha de rubricas:\n\n{e}")

def save_imported_payroll_codes(code_list):
    """Recebe uma lista de rubricas e salva apenas as 'Novas' no Firestore."""
    db = firestore.client()
    try:
        added_count = 0
        batch = db.batch()
        codes_ref = db.collection('payroll_codes')

        for code_item in code_list:
            if code_item['status'] == "Novo":
                new_doc_ref = codes_ref.document()
                batch.set(new_doc_ref, {
                    'code': code_item['code'],
                    'name': code_item['name'],
                    'value_type': code_item['value_type'],
                    'calculation_base': code_item.get('calculation_base', 'Valor Informado'),
                    'calculation_factor': code_item.get('calculation_factor', 1.0)
                })
                added_count += 1
        
        batch.commit()
        
        return (True, f"{added_count} nova(s) rubrica(s) foram importadas com sucesso.")
    except Exception as e:
        return (False, f"Ocorreu um erro ao salvar os dados no banco de dados:\n\n{e}")


# --- LÓGICA PARA LANÇAMENTOS E GERAÇÃO DO ARQUIVO TXT ---

def format_value(value):
    """Formata o valor para o arquivo TXT com 15 posições, preenchido com zeros."""
    try:
        numeric_value = int(float(str(value).replace(',', '.')) * 100)
        return f"{numeric_value:015d}"
    except (ValueError, TypeError):
        return "0" * 15

def calculate_payroll_value(salary, base, factor, quantity, monthly_hours=220.0):
    """
    Calcula o valor final de uma rubrica com base em seus parâmetros.
    Retorna o valor calculado e a memória de cálculo como string.
    """
    try:
        if base == "Baseado no Salário-Hora":
            if monthly_hours <= 0: return 0.0, "Divisor de horas inválido."
            salary_hour = salary / monthly_hours
            calculated_value = salary_hour * quantity * factor
            memo = f"(R$ {salary:.2f} / {monthly_hours}h) * {quantity}h * {factor} = R$ {calculated_value:.2f}"
            return calculated_value, memo
        elif base == "Percentual sobre o Salário":
            calculated_value = salary * (quantity / 100.0) * factor
            memo = f"R$ {salary:.2f} * {quantity}% * {factor} = R$ {calculated_value:.2f}"
            return calculated_value, memo
        else:
            return quantity, "Cálculo não aplicável."
    except Exception as e:
        return 0.0, f"Erro no cálculo: {e}"

def generate_import_file(company_code, calc_type, competence, all_launches, save_path):
    """
    Organiza os lançamentos e gera o arquivo TXT final no formato de 50 colunas.
    """
    try:
        launches_by_employee = {}
        for launch in all_launches:
            emp_code = launch['employee_code']
            if emp_code not in launches_by_employee:
                launches_by_employee[emp_code] = []
            launches_by_employee[emp_code].append(launch)

        file_content = []
        for emp_code, launches in launches_by_employee.items():
            for launch in launches:
                # --- [ALTERADO] A competência agora é formatada para YYYYMM ---
                clean_competence = ""
                if competence and "/" in competence:
                    parts = competence.split('/')
                    if len(parts) == 2:
                        mm, yyyy = parts
                        clean_competence = f"{yyyy}{mm}"
                
                # Se o formato for inesperado, usa como está, removendo a barra
                if not clean_competence:
                    clean_competence = competence.replace('/', '')

                line = (
                    f"{int(company_code):05d}"
                    f"{int(emp_code):06d}"
                    f"{int(launch['payroll_code']['code']):05d}"
                    f"{format_value(launch['value'])}"
                    f"{clean_competence:6}"
                    f"{int(calc_type):02d}"
                    f"{'': <11}" # 11 espaços em branco para completar 50 colunas
                )
                file_content.append(line)
        
        with open(save_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(file_content))
            
        return (True, f"Arquivo de importação gerado com sucesso em:\n{save_path}")

    except Exception as e:
        logging.error(f"Erro ao gerar arquivo TXT: {e}", exc_info=True)
        return (False, f"Ocorreu um erro inesperado ao gerar o arquivo TXT:\n\n{e}")