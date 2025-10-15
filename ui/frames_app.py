# ===================================================================================
# --- ARQUIVO: ui/frames_app.py ---
# (Contém as telas principais da aplicação após o login)
# ===================================================================================

import os
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext, simpledialog
from collections import defaultdict
from PIL import Image, ImageTk, ImageEnhance, ImageOps
import ttkbootstrap as ttk
from ttkbootstrap.tooltip import ToolTip
from datetime import datetime
import logging
from openpyxl import Workbook
import threading

import core_logic
import auth_logic
from .dialogs_flow import DashboardWindow, PreviewWindow


class HomeFrame(ttk.Frame):
    def __init__(self, parent, controller, *args): 
        super().__init__(parent)
        self.controller = controller
        self.user_display_var = tk.StringVar()
        self.columnconfigure(0, weight=1); self.columnconfigure(1, weight=1); self.rowconfigure(1, weight=1)
        user_info_frame = ttk.Frame(self); user_info_frame.place(relx=0.99, rely=0.01, anchor='ne')
        ttk.Label(user_info_frame, textvariable=self.user_display_var, bootstyle="secondary").pack(side="left", padx=(0, 10))
        ttk.Button(user_info_frame, text="Trocar de Usuário", command=lambda: self.controller.logout(), bootstyle="light-outline").pack(side="left")
        title_frame = ttk.Frame(self); title_frame.grid(row=0, column=0, columnspan=2, pady=(40, 10), sticky='ew')
        title_frame.columnconfigure(0, weight=1)
        ttk.Label(title_frame, text="Fiscal Flow - Customs Trading", font=("Helvetica", 24, "bold")).pack()
        left_container = ttk.Frame(self); left_container.grid(row=1, column=0, sticky='nsew')
        buttons_frame = ttk.Frame(left_container); buttons_frame.pack(expand=True)
        style = self.controller.app_style; style.configure('Large.TButton', font=('Helvetica', 12))
        self.btn_extracao = ttk.Button(buttons_frame, text="Ferramentas de Extração", style='Large.TButton', width=40, bootstyle="primary")
        self.btn_extracao.config(command=lambda: controller.pulse_and_navigate(self.btn_extracao, "ExtractionToolsFrame")); self.btn_extracao.pack(pady=10, ipady=5)
        self.btn_dp = ttk.Button(buttons_frame, text="Departamento Pessoal", style='Large.TButton', width=40, bootstyle="primary")
        self.btn_dp.config(command=lambda: controller.pulse_and_navigate(self.btn_dp, "DPMainFrame")); self.btn_dp.pack(pady=10, ipady=5)
        self.btn_contabil = ttk.Button(buttons_frame, text="Departamento Contábil", style='Large.TButton', width=40, state="disabled", bootstyle="secondary"); self.btn_contabil.pack(pady=10, ipady=5)
        self.btn_fiscal = ttk.Button(buttons_frame, text="Departamento Fiscal", style='Large.TButton', width=40, state="disabled", bootstyle="secondary"); self.btn_fiscal.pack(pady=10, ipady=5)
        self.right_container = ttk.Frame(self); self.right_container.grid(row=1, column=1, sticky='nsew')
        self.watermark_label = ttk.Label(self.right_container); self.text_label = ttk.Label(self.right_container)
        support_icons_frame = ttk.Frame(self)
        support_icons_frame.place(relx=1.0, rely=1.0, x=-15, y=-15, anchor='se')
        self.support_it_btn = ttk.Button(support_icons_frame, command=lambda: self.controller.show_support_tickets('it'), bootstyle="link")
        self.support_it_btn.pack(pady=2)
        self.support_system_btn = ttk.Button(support_icons_frame, command=lambda: self.controller.show_support_tickets('developer'), bootstyle="link")
        self.support_system_btn.pack(pady=2)
        ToolTip(self.support_it_btn, text="Acessar Suporte de T.I.")
        ToolTip(self.support_system_btn, text="Acessar Suporte do Sistema")
        self.update_watermarks()
    def update_user_display(self, username):
        self.user_display_var.set(f"Logado como: {username}")
    def on_frame_activated(self):
        user = self.controller.current_user
        if not user: return
        user_level = user.get('level')
        if user_level in ['Admin', 'Desenvolvedor']:
            self.btn_extracao.config(state="normal"); self.btn_dp.config(state="normal"); return
        user_sector_ids = user.get('sector_ids', [])
        all_sectors, error = self.controller.get_sectors()
        if error: self.btn_extracao.config(state="disabled"); self.btn_dp.config(state="disabled"); return
        sector_id_map = {s['id']: s['name'] for s in all_sectors}
        user_sector_names = {sector_id_map.get(sid) for sid in user_sector_ids}
        ferramentas_permitidas = {'Contábil', 'Pessoal', 'Fiscal'}
        tem_acesso = not ferramentas_permitidas.isdisjoint(user_sector_names)
        novo_estado = "normal" if tem_acesso else "disabled"
        self.btn_extracao.config(state=novo_estado); self.btn_dp.config(state=novo_estado)
    def _create_themed_icon(self, filename, size=(32, 32)):
        theme_name = self.controller.app_style.theme.name; color_hex = self.controller.app_style.colors.get('primary')
        cache_key = f"{filename}_{theme_name}"
        if cache_key in self.controller.image_cache: return self.controller.image_cache[cache_key]
        try:
            base_image = Image.open(self.controller.resource_path(filename)).convert("RGBA")
            color_image = Image.new("RGBA", base_image.size, color=color_hex)
            final_image = Image.new("RGBA", base_image.size, (0, 0, 0, 0))
            final_image = Image.composite(color_image, final_image, base_image)
            final_image = final_image.resize(size, Image.LANCZOS)
            photo_image = ImageTk.PhotoImage(final_image)
            self.controller.image_cache[cache_key] = photo_image
            return photo_image
        except Exception as e:
            logging.warning(f"Não foi possível carregar/colorir o ícone '{filename}': {e}"); return None
    def update_watermarks(self):
        self.support_it_icon = self._create_themed_icon("icon_support_it.png") 
        if self.support_it_icon: self.support_it_btn.config(image=self.support_it_icon); self.support_it_btn.image = self.support_it_icon
        self.support_system_icon = self._create_themed_icon("icon_support_system.png")
        if self.support_system_icon: self.support_system_btn.config(image=self.support_system_icon); self.support_system_btn.image = self.support_system_icon
        theme_type = self.controller.app_style.theme.type
        logo_file = 'logo_light.png' if theme_type == 'dark' else 'logo_dark.png'
        text_file = 'logo_text_light.png' if theme_type == 'dark' else 'logo_text_dark.png'
        try:
            if logo_file not in self.controller.image_cache:
                watermark_pil = Image.open(self.controller.resource_path(logo_file)).resize((450, 450), Image.LANCZOS).convert("RGBA")
                enhancer = ImageEnhance.Brightness(watermark_pil); watermark_pil = enhancer.enhance(0.4)
                alpha = watermark_pil.getchannel('A'); alpha = alpha.point(lambda p: p // 6); watermark_pil.putalpha(alpha)
                self.controller.image_cache[logo_file] = ImageTk.PhotoImage(watermark_pil)
            self.watermark_image = self.controller.image_cache[logo_file]
            self.watermark_label.config(image=self.watermark_image); self.watermark_label.pack(expand=True)
            if text_file not in self.controller.image_cache:
                text_pil = Image.open(self.controller.resource_path(text_file)).convert("RGBA")
                alpha_text = text_pil.getchannel('A'); alpha_text = alpha_text.point(lambda p: p // 6); text_pil.putalpha(alpha_text)
                self.controller.image_cache[text_file] = ImageTk.PhotoImage(text_pil)
            self.text_image = self.controller.image_cache[text_file]
            self.text_label.config(image=self.text_image); self.text_label.place(relx=0.5, rely=0.75, anchor='center')
        except Exception as e: logging.warning(f"Não foi possível carregar as logos do menu: {e}")

class ExtractionToolsFrame(ttk.Frame):
    def __init__(self, parent, controller, *args):
        super().__init__(parent); self.controller = controller; self.columnconfigure(0, weight=1); self.rowconfigure(0, weight=1)
        main_container = ttk.Frame(self); main_container.pack(expand=True)
        ttk.Label(main_container, text="Ferramentas de Extração", font=("Helvetica", 24, "bold")).pack(pady=(0, 40))
        buttons_frame = ttk.Frame(main_container); buttons_frame.pack()
        style = self.controller.app_style; style.configure('Large.TButton', font=('Helvetica', 12))
        btn_nfe = ttk.Button(buttons_frame, text="Extrator de NFe (XML)", style='Large.TButton', width=40)
        btn_nfe.config(command=lambda: controller.pulse_and_navigate(btn_nfe, "NFeToolFrame")); btn_nfe.pack(pady=10, ipady=5)
        ttk.Button(buttons_frame, text="(Futura Ferramenta de Extração)", style='Large.TButton', width=40, state="disabled").pack(pady=10, ipady=5)
        ttk.Button(self, text="< Voltar ao Menu Principal", command=lambda: controller.show_frame("HomeFrame")).place(relx=0.01, rely=0.98, anchor='sw')

class NFeToolFrame(ttk.Frame):
    def __init__(self, parent, controller, *args):
        super().__init__(parent); self.controller = controller; self.dados_extraidos_em_memoria = []
        self.columnconfigure(0, weight=1); self.rowconfigure(3, weight=1); self.create_widgets()

    def create_widgets(self):
        self.watermark_label_tool = tk.Label(self, bd=0)
        self.update_watermarks() 
        top_frame = ttk.Frame(self); top_frame.grid(row=0, column=0, pady=(10,5), sticky='ew'); top_frame.columnconfigure(1, weight=1)
        btn_voltar = ttk.Button(top_frame, text="< Voltar", command=lambda: self.controller.pulse_and_navigate(btn_voltar, "ExtractionToolsFrame")); btn_voltar.grid(row=0, column=0, padx=10, sticky='w')
        ttk.Label(top_frame, text="Extrator de NFe", font=("Helvetica", 18, "bold")).grid(row=0, column=1)
        path_frame = ttk.Frame(self, padding="10"); path_frame.grid(row=1, column=0, sticky='ew'); path_frame.columnconfigure(1, weight=1)
        ttk.Label(path_frame, text="Pasta XML:").grid(row=0, column=0, padx=(0,5), sticky='w')
        self.path_entry = ttk.Entry(path_frame); self.path_entry.grid(row=0, column=1, padx=5, sticky='ew')
        if self.controller.default_xml_path: self.path_entry.insert(0, self.controller.default_xml_path)
        ttk.Button(path_frame, text="Selecionar", command=self.selecionar_pasta_origem).grid(row=0, column=2, padx=(5,0))
        self.progress_bar = ttk.Progressbar(self, mode='determinate'); self.progress_bar.grid(row=2, column=0, sticky='ew', padx=10, pady=5)
        ttk.Frame(self).grid(row=3, column=0) 
        self.status_var = tk.StringVar(value="Pronto.")
        status_frame = ttk.Frame(self, padding=(10, 5)); status_frame.grid(row=4, column=0, sticky='ew'); status_frame.columnconfigure(0, weight=1)
        ttk.Label(status_frame, textvariable=self.status_var, font=("Helvetica", 8)).grid(row=0, column=0, sticky='w')
        controls_frame = ttk.Frame(self, padding="10"); controls_frame.grid(row=5, column=0, pady=5, sticky='ew'); controls_frame.columnconfigure(0, weight=1)
        ttk.Button(controls_frame, text="Importar e Analisar XMLs", command=self.extrair_dados_e_analisar, width=30).grid(row=0, column=0, pady=5, sticky='ew', padx=5)
        ttk.Button(controls_frame, text="Gerar Diagnóstico de Processos", command=self.gerar_diagnostico, bootstyle="info").grid(row=1, column=0, pady=5, padx=5, sticky='ew')
        ttk.Button(controls_frame, text="Limpar", command=self.limpar_dados_e_interface).grid(row=2, column=0, pady=5, padx=5, sticky='ew')
        signature_frame = ttk.Frame(self, padding=(10, 5)); signature_frame.grid(row=6, column=0, sticky='ew'); signature_frame.columnconfigure(0, weight=1)
        ttk.Label(signature_frame, text="Desenvolvido por Bruno Silva - Analista Contábil", font=("Helvetica", 8)).pack()

    def update_watermarks(self):
        theme_type = self.controller.app_style.theme.type; logo_file = 'logo_light.png' if theme_type == 'dark' else 'logo_dark.png'; cache_key = f"{logo_file}_tool"
        try:
            if cache_key not in self.controller.image_cache:
                watermark_pil = Image.open(self.controller.resource_path(logo_file)).resize((350, 350), Image.LANCZOS).convert("RGBA")
                enhancer = ImageEnhance.Brightness(watermark_pil); watermark_pil = enhancer.enhance(0.4)
                alpha = watermark_pil.getchannel('A'); alpha = alpha.point(lambda p: p // 6); watermark_pil.putalpha(alpha)
                self.controller.image_cache[cache_key] = ImageTk.PhotoImage(watermark_pil)
            self.watermark_image_tool = self.controller.image_cache[cache_key]
            self.watermark_label_tool.config(image=self.watermark_image_tool); self.watermark_label_tool.place(relx=0.5, rely=0.45, anchor='center')
        except Exception as e: logging.warning(f"Não foi possível carregar a logo para a marca d'água: {e}")
    
    # --- [FUNÇÃO RESTAURADA] ---
    def update_status(self, message):
        self.status_var.set(message)
        logging.info(message)
        self.controller.update_idletasks()

    def gerar_diagnostico(self):
        if not self.dados_extraidos_em_memoria:
            self.extrair_dados_e_analisar(run_dashboard=False)
            if not self.dados_extraidos_em_memoria: return
        caminho_saida = filedialog.asksaveasfilename(title="Salvar Diagnóstico", initialfile="diagnostico_processos.txt", defaultextension=".txt", filetypes=[("Arquivos de Texto", "*.txt")])
        if not caminho_saida: return
        success, message = core_logic.gerar_diagnostico_processos(self.dados_extraidos_em_memoria, caminho_saida)
        if success:
            messagebox.showinfo("Sucesso", message, parent=self)
            if messagebox.askyesno("Abrir Arquivo", "Deseja abrir o diagnóstico?"): os.startfile(caminho_saida)
        else: messagebox.showerror("Erro", message, parent=self)
        
    def extrair_dados_e_analisar(self, run_dashboard=True):
        try:
            self.dados_extraidos_em_memoria = []; self.progress_bar['value'] = 0; self.update_status("Iniciando extração...")
            pasta_xml = self.path_entry.get()
            if not pasta_xml or not os.path.isdir(pasta_xml): self.update_status("Erro: Pasta inválida."); messagebox.showerror("Erro", "Caminho inválido."); return
            arquivos_xml = [os.path.join(dp, f) for dp, _, fn in os.walk(pasta_xml) for f in fn if f.lower().endswith('.xml')]
            total = len(arquivos_xml); self.progress_bar['maximum'] = total; self.update_status(f"Encontrados {total} arquivos XML.")
            if total == 0: messagebox.showinfo("Aviso", "Nenhum XML encontrado."); self.update_status("Pronto."); return
            erros = []
            for i, arq_path in enumerate(arquivos_xml):
                self.update_status(f"Processando {i + 1}/{total}: {os.path.basename(arq_path)}"); self.progress_bar['value'] = i + 1
                dados = core_logic.extrair_dados_nf(arq_path)
                if dados: self.dados_extraidos_em_memoria.append(dados)
                else: erros.append(os.path.basename(arq_path))
            if erros: messagebox.showwarning("Aviso", "Falha ao processar:\n" + "\n".join(erros))
            self.update_status(f"Extração concluída. Calculando resumo...")
            if run_dashboard:
                dashboard_data = core_logic.calcular_dados_dashboard(self.dados_extraidos_em_memoria)
                DashboardWindow(self.controller, dashboard_data); self.update_status(f"Análise concluída.")
        except Exception as e:
            logging.error("Erro na extração.", exc_info=True); messagebox.showerror("Erro Crítico", f"Ocorreu um erro:\n{e}"); self.update_status("Erro crítico.")
            
    def salvar_dados_basicos(self):
        if not self.dados_extraidos_em_memoria: messagebox.showerror("Erro", "Nenhum dado extraído."); return
        filename = self.controller.output_filename_pattern.format(data=datetime.now().strftime('%Y-%m-%d'))
        initial_dir = self.controller.default_output_path or os.path.expanduser("~")
        caminho_excel = filedialog.asksaveasfilename(initialdir=initial_dir, initialfile=filename, defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not caminho_excel: return
        self.update_status("Gerando Excel Básico..."); wb = Workbook()
        saida_ws, entrada_ws = wb.create_sheet("Saída"), wb.create_sheet("Entrada"); del wb['Sheet']
        cabecalhos = ['Arquivo', 'Número da NF', 'Cliente', 'Data de Emissão', 'Sistema Emissor', 'Valor Total', 'Status da NF']
        core_logic.setup_headers(entrada_ws, cabecalhos); core_logic.setup_headers(saida_ws, cabecalhos)
        row_in, row_out = 2, 2
        for dados in self.dados_extraidos_em_memoria:
            if dados['tipo_nota'] == "Saída": core_logic.write_data_to_excel(saida_ws, row_out, dados, cabecalhos); row_out += 1
            else: core_logic.write_data_to_excel(entrada_ws, row_in, dados, cabecalhos); row_in += 1
        core_logic.add_totals_row(entrada_ws, cabecalhos); core_logic.add_totals_row(saida_ws, cabecalhos); self.salvar_arquivo(wb, caminho_excel)
    def salvar_planilha_completa(self, contagens_confirmadas):
        if not self.dados_extraidos_em_memoria: messagebox.showerror("Erro", "Nenhum dado extraído."); return
        filename = self.controller.output_filename_pattern.format(data=datetime.now().strftime('%Y-%m-%d'))
        initial_dir = self.controller.default_output_path or os.path.expanduser("~")
        caminho_excel = filedialog.asksaveasfilename(initialdir=initial_dir, initialfile=filename, defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not caminho_excel: return
        self.update_status("Gerando Planilha Completa..."); wb = Workbook()
        saida_ws, entrada_ws = wb.create_sheet("Saída"), wb.create_sheet("Entrada")
        servico_ws, pendentes_ws = wb.create_sheet("Serviços Autorizados"), wb.create_sheet("Serviços Pendentes"); del wb['Sheet']
        cabecalhos_completos = ['Arquivo', 'Número da NF', 'Status da NF', 'Data de Emissão', 'Cliente', 'CNPJ/CPF Dest.', 'Razão Social Dest.', 'UF Destino', 'Nome do Processo', 'CFOP', 'Valor Total dos Produtos', 'Valor II', 'Valor ICMS', 'Valor IPI', 'Valor PIS', 'Valor COFINS', 'Outras Despesas', 'Valor AFRMM', 'Frete Nacional', 'Valor Total', 'Sistema Emissor']
        cabecalhos_servico = ['Número da NF', 'Cliente', 'Data de Emissão', 'Sistema Emissor', 'Nome do Processo', 'Valor Serviço Trading']
        cabecalhos_pendentes = cabecalhos_servico + ['Qtde Encontrada', 'Qtde Esperada', 'Qtde a Emitir']
        core_logic.setup_headers(entrada_ws, cabecalhos_completos); core_logic.setup_headers(saida_ws, cabecalhos_completos)
        core_logic.setup_headers(servico_ws, cabecalhos_servico); core_logic.setup_headers(pendentes_ws, cabecalhos_pendentes)
        row_in, row_out, row_serv, row_pend = 2, 2, 2, 2
        processos_saida = defaultdict(list)
        for dados in self.dados_extraidos_em_memoria:
            if dados['tipo_nota'] == 'Saída': processos_saida[dados.get('processo_normalizado', 'N/A')].append(dados)
        for dados in self.dados_extraidos_em_memoria:
            if dados['tipo_nota'] == "Saída": core_logic.write_data_to_excel(saida_ws, row_out, dados, cabecalhos_completos); row_out += 1
            else: core_logic.write_data_to_excel(entrada_ws, row_in, dados, cabecalhos_completos); row_in += 1
            if dados.get('valor_servico_trading', 0.0) > 0.0 and dados['tipo_nota'] == 'Entrada':
                linha_serv = [dados['numero_nf'], dados['nome_cliente'], dados['data_emissao'], dados.get('sistema_emissor'), dados['nome_processo'], dados['valor_servico_trading']]
                processo = dados.get('processo_normalizado', 'N/A')
                nfs_encontradas = len(processos_saida.get(processo, [])); esperada = contagens_confirmadas.get(processo, {}).get('esperado')
                auth = (esperada is not None and nfs_encontradas == esperada) or (nfs_encontradas > 0 and processo not in contagens_confirmadas)
                if auth:
                    for col, val in enumerate(linha_serv, 1): servico_ws.cell(row=row_serv, column=col, value=val).number_format = 'R$ #,##0.00' if col == 6 else None
                    row_serv += 1
                else:
                    a_emitir = esperada - nfs_encontradas if esperada is not None else 'N/A'
                    linha_pend = linha_serv + [nfs_encontradas, esperada, a_emitir]
                    for col, val in enumerate(linha_pend, 1): pendentes_ws.cell(row=row_pend, column=col, value=val).number_format = 'R$ #,##0.00' if col == 6 else None
                    row_pend += 1
        core_logic.add_totals_row(entrada_ws, cabecalhos_completos); core_logic.add_totals_row(saida_ws, cabecalhos_completos)
        core_logic.add_totals_row(servico_ws, cabecalhos_servico); core_logic.add_totals_row(pendentes_ws, cabecalhos_pendentes)
        self.salvar_arquivo(wb, caminho_excel)
    def salvar_arquivo(self, wb, caminho):
        try:
            wb.save(caminho); self.update_status(f"Arquivo salvo!");
            if self.controller.ask_to_open_excel and messagebox.askyesno("Abrir", "Deseja abrir o arquivo?"): os.startfile(caminho)
        except Exception as e:
            self.update_status(f"Erro ao salvar."); logging.error(f"Erro ao salvar: {e}", exc_info=True); messagebox.showerror("Erro", f"Erro ao salvar o arquivo: {e}")
    def limpar_dados_e_interface(self):
        if messagebox.askyesno("Confirmação", "Limpar todos os dados?"):
            self.dados_extraidos_em_memoria.clear(); self.path_entry.delete(0, tk.END); self.progress_bar.config(value=0)
            self.update_status("Pronto."); logging.info("Interface limpa.")
    def selecionar_pasta_origem(self):
        initial_dir = self.controller.default_xml_path or os.path.expanduser("~")
        folder_path = filedialog.askdirectory(title="Selecione a pasta XML", initialdir=initial_dir)
        if folder_path: self.path_entry.delete(0, tk.END); self.path_entry.insert(0, folder_path); self.update_status(f"Pasta selecionada: {folder_path}")
    def show_preview_window(self, contagens_confirmadas):
        if not self.dados_extraidos_em_memoria: messagebox.showinfo("Aviso", "Não há dados para pré-visualizar.", parent=self.controller); return
        PreviewWindow(controller=self.controller, all_extracted_data=self.dados_extraidos_em_memoria, contagens_confirmadas=contagens_confirmadas)